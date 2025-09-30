"""
Generador de reportes en Excel para los resultados de las pruebas E2E
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os


class ExcelReporter:
    """Clase para generar reportes en Excel de los resultados de las pruebas"""

    def __init__(self, filename="reporte_pruebas.xlsx"):
        self.filename = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Resultados de Pruebas"
        self.current_row = 1
        self._setup_headers()

    def _setup_headers(self):
        """Configura los encabezados de las columnas"""
        headers = [
            "ID Caso",
            "Historia de Usuario",
            "Nombre del Test",
            "Descripción",
            "Estado",
            "Fecha Ejecución",
            "Duración (seg)",
            "Error/Observaciones"
        ]

        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = border

        # Ajustar anchos de columna
        column_widths = [12, 30, 40, 50, 15, 20, 15, 50]
        for col, width in enumerate(column_widths, start=1):
            self.sheet.column_dimensions[chr(64 + col)].width = width

        self.current_row = 2

    def add_test_result(self, test_id, historia, test_name, description, status, duration=0, error_msg=""):
        """
        Agrega un resultado de test al reporte

        Args:
            test_id (str): ID del caso de prueba (ej: CP-01)
            historia (str): Nombre de la historia de usuario
            test_name (str): Nombre del test
            description (str): Descripción del test
            status (str): Estado del test (PASSED, FAILED, SKIPPED, ERROR)
            duration (float): Duración en segundos
            error_msg (str): Mensaje de error si aplica
        """
        row = self.current_row

        # Valores
        self.sheet.cell(row=row, column=1).value = test_id
        self.sheet.cell(row=row, column=2).value = historia
        self.sheet.cell(row=row, column=3).value = test_name
        self.sheet.cell(row=row, column=4).value = description
        self.sheet.cell(row=row, column=5).value = status
        self.sheet.cell(row=row, column=6).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(row=row, column=7).value = round(duration, 2)
        self.sheet.cell(row=row, column=8).value = error_msg

        # Estilo según estado
        status_colors = {
            "PASSED": "C6EFCE",  # Verde claro
            "FAILED": "FFC7CE",  # Rojo claro
            "SKIPPED": "FFEB9C",  # Amarillo claro
            "ERROR": "FF9999"    # Rojo más oscuro
        }

        status_cell = self.sheet.cell(row=row, column=5)
        if status in status_colors:
            status_cell.fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type="solid"
            )
            status_cell.font = Font(bold=True)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Bordes
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, 9):
            cell = self.sheet.cell(row=row, column=col)
            cell.border = border
            if col in [1, 5, 6, 7]:  # Centrar algunas columnas
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        self.current_row += 1

    def add_summary(self):
        """Agrega un resumen al final del reporte"""
        # Dejar una fila en blanco
        self.current_row += 1

        # Título de resumen
        summary_row = self.current_row
        self.sheet.cell(row=summary_row, column=1).value = "RESUMEN"
        self.sheet.cell(row=summary_row, column=1).font = Font(bold=True, size=14)

        # Contar estados
        passed = 0
        failed = 0
        skipped = 0
        error = 0

        for row in range(2, summary_row - 1):
            status = self.sheet.cell(row=row, column=5).value
            if status == "PASSED":
                passed += 1
            elif status == "FAILED":
                failed += 1
            elif status == "SKIPPED":
                skipped += 1
            elif status == "ERROR":
                error += 1

        total = passed + failed + skipped + error

        # Agregar estadísticas
        self.current_row += 1
        self.sheet.cell(row=self.current_row, column=1).value = "Total de Pruebas:"
        self.sheet.cell(row=self.current_row, column=2).value = total
        self.current_row += 1

        self.sheet.cell(row=self.current_row, column=1).value = "Exitosas (PASSED):"
        self.sheet.cell(row=self.current_row, column=2).value = passed
        self.sheet.cell(row=self.current_row, column=2).fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        self.current_row += 1

        self.sheet.cell(row=self.current_row, column=1).value = "Fallidas (FAILED):"
        self.sheet.cell(row=self.current_row, column=2).value = failed
        self.sheet.cell(row=self.current_row, column=2).fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        self.current_row += 1

        self.sheet.cell(row=self.current_row, column=1).value = "Omitidas (SKIPPED):"
        self.sheet.cell(row=self.current_row, column=2).value = skipped
        self.sheet.cell(row=self.current_row, column=2).fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        self.current_row += 1

        if total > 0:
            porcentaje_exito = (passed / total) * 100
            self.current_row += 1
            self.sheet.cell(row=self.current_row, column=1).value = "Porcentaje de Éxito:"
            self.sheet.cell(row=self.current_row, column=2).value = f"{porcentaje_exito:.2f}%"
            self.sheet.cell(row=self.current_row, column=2).font = Font(bold=True)

    def save(self, directory="reports"):
        """
        Guarda el reporte en un archivo Excel

        Args:
            directory (str): Directorio donde guardar el reporte
        """
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Agregar timestamp al nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_with_timestamp = f"reporte_pruebas_{timestamp}.xlsx"
        filepath = os.path.join(directory, filename_with_timestamp)

        self.add_summary()
        self.workbook.save(filepath)
        print(f"\n✓ Reporte Excel generado: {filepath}")
        return filepath
